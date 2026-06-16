"""
Download and parse Nature Communications 2020 supplementary data
(Primate arterial aging scRNA-seq, Young vs Old DEG lists)
Then cross-reference with Ferro-aging gene set
"""
import urllib.request
import openpyxl
import pandas as pd
import os

OUTDIR = "C:/Users/lidaf/WorkBuddy/2026-06-11-19-29-35/results"
SUPP_URLS = [
    ("MOESM7", "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-020-15997-0/MediaObjects/41467_2020_15997_MOESM7_ESM.xlsx"),
    ("MOESM8", "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-020-15997-0/MediaObjects/41467_2020_15997_MOESM8_ESM.xlsx"),
    ("MOESM9", "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-020-15997-0/MediaObjects/41467_2020_15997_MOESM9_ESM.xlsx"),
    ("MOESM10", "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-020-15997-0/MediaObjects/41467_2020_15997_MOESM10_ESM.xlsx"),
]

# Load ferro-aging gene set
fa_df = pd.read_csv("C:/Users/lidaf/WorkBuddy/2026-06-11-19-29-35/ferro_aging_geneset.csv")
fa_genes = set(fa_df['gene_symbol'].str.upper())

print("=" * 60)
print("Primate Arterial Aging (Nat Commun 2020) - Ferro-aging Cross-Reference")
print("=" * 60)

for name, url in SUPP_URLS:
    fname = os.path.join(OUTDIR, f"primate_aging_{name}.xlsx")
    print(f"\nDownloading {name}...")
    try:
        urllib.request.urlretrieve(url, fname)
        wb = openpyxl.load_workbook(fname, data_only=True, read_only=True)
        sheets = wb.sheetnames
        print(f"  Sheets: {sheets}")
        
        for sheet_name in sheets:
            ws = wb[sheet_name]
            # Read first 5 rows to understand structure
            print(f"  Sheet '{sheet_name}' - {ws.max_row} rows x {ws.max_column} cols")
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                vals = [str(v)[:50] if v is not None else '' for v in row]
                print(f"    Row {row_idx}: {vals}")
            
            # Try to find gene names and cross-reference with FA genes
            # Look for gene symbol column
            gene_col_idx = None
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                header = str(ws.cell(1, col_idx).value or '').lower()
                if 'gene' in header or 'symbol' in header or 'name' in header:
                    gene_col_idx = col_idx
                    print(f"  → Found gene column at col {col_idx}: '{ws.cell(1, col_idx).value}'")
                    break
            
            if gene_col_idx:
                # Extract all genes from this sheet
                all_genes = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    gene = str(row[gene_col_idx - 1]).strip().upper() if row[gene_col_idx - 1] else ''
                    if gene and gene != 'NONE' and len(gene) < 20:
                        all_genes.append(gene)
                
                fa_hits = set(all_genes) & fa_genes
                print(f"  → Total genes in sheet: {len(all_genes)}")
                print(f"  → Ferro-aging genes found: {len(fa_hits)}")
                if fa_hits:
                    print(f"  → FA genes: {sorted(fa_hits)}")
        
        wb.close()
    except Exception as e:
        print(f"  Error: {e}")

print("\nDone!")
